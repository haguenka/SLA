import streamlit as st
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.drawing.image import Image as OpenpyxlImage
from io import BytesIO
import matplotlib.pyplot as plt

# Function to apply conditional formatting for colors
def apply_conditional_formatting(output_file):
    wb = load_workbook(output_file)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

        # Apply conditional formatting based on the value of 'SLA_STATUS'
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=ws.max_column, max_col=ws.max_column):
            sla_status_cell = row[0]

            if sla_status_cell.value == 'SLA FORA DO PERÍODO':
                for cell in row:
                    cell.fill = red_fill
            elif sla_status_cell.value == 'SLA DENTRO DO PERÍODO':
                for cell in row:
                    cell.fill = green_fill

    wb.save(output_file)

# Streamlit app
def main():
    st.title("SLA Analysis Dashboard")

    # File upload
    uploaded_file = st.file_uploader("Upload an Excel file for SLA Analysis", type=['xlsx'])
    if uploaded_file is not None:
        output_folder = st.text_input("Enter the output folder path:")
        if st.button("Run SLA Analysis"):
            try:
                # Load the Excel file
                df = pd.read_excel(uploaded_file)

                # Filter by GRUPO to include only 'GRUPO TOMOGRAFIA' and 'GRUPO RESSONANCIA MAGNETICA'
                allowed_groups = ['GRUPO TOMOGRAFIA', 'GRUPO RESSONÂNCIA MAGNÉTICA']
                df = df[df['GRUPO'].isin(allowed_groups)]

                # Parse the relevant datetime columns explicitly with dayfirst=True
                df['STATUS_ALAUDAR'] = pd.to_datetime(df['STATUS_ALAUDAR'], dayfirst=True, errors='coerce')
                df['STATUS_PRELIMINAR'] = pd.to_datetime(df['STATUS_PRELIMINAR'], dayfirst=True, errors='coerce')
                df['STATUS_APROVADO'] = pd.to_datetime(df['STATUS_APROVADO'], dayfirst=True, errors='coerce')

                # Eliminate rows where both STATUS_PRELIMINAR and STATUS_APROVADO are null
                df = df.dropna(subset=['STATUS_PRELIMINAR', 'STATUS_APROVADO'], how='all')

                # Check if 'UNIDADE' and 'TIPO_ATENDIMENTO' exist
                if 'UNIDADE' not in df.columns or 'TIPO_ATENDIMENTO' not in df.columns:
                    st.error("'UNIDADE' or 'TIPO_ATENDIMENTO' column not found.")
                    return

                # Calculate DELTA_TIME using STATUS_PRELIMINAR if it exists, otherwise use STATUS_APROVADO
                df['DELTA_TIME'] = (df['STATUS_PRELIMINAR'].fillna(df['STATUS_APROVADO']) - df['STATUS_ALAUDAR']).dt.total_seconds() / 3600  # in hours

                # Define the conditions for SLA violations
                conditions = [
                    (df['TIPO_ATENDIMENTO'] == 'Pronto Atendimento') & (df['DELTA_TIME'] > 1),
                    (df['TIPO_ATENDIMENTO'] == 'Internado') & (df['DELTA_TIME'] > 24),
                    (df['TIPO_ATENDIMENTO'] == 'Externo') & (df['DELTA_TIME'] > 72)
                ]

                # Set the default SLA status and apply conditions
                df['SLA_STATUS'] = 'SLA DENTRO DO PERÍODO'
                df.loc[conditions[0], 'SLA_STATUS'] = 'SLA FORA DO PERÍODO'
                df.loc[conditions[1], 'SLA_STATUS'] = 'SLA FORA DO PERÍODO'
                df.loc[conditions[2], 'SLA_STATUS'] = 'SLA FORA DO PERÍODO'

                # Select only relevant columns
                selected_columns = [
                    'SAME', 'NOME_PACIENTE', 'GRUPO', 'DESCRICAO_PROCEDIMENTO', 'MEDICO_LAUDO_DEFINITIVO',
                    'UNIDADE', 'TIPO_ATENDIMENTO', 'STATUS_ALAUDAR', 'STATUS_PRELIMINAR', 'STATUS_APROVADO', 'DELTA_TIME', 'SLA_STATUS'
                ]

                df_selected = df[selected_columns]

                # Split the data by 'UNIDADE' and 'TIPO_ATENDIMENTO' and save them to separate sheets
                output_file = os.path.join(output_folder, "sla_analysis_output.xlsx")
                with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                    # First group by 'UNIDADE' and then by 'TIPO_ATENDIMENTO'
                    for (unidade, tipo_atendimento), group in df_selected.groupby(['UNIDADE', 'TIPO_ATENDIMENTO']):
                        if not group.empty:
                            sheet_name = f"{unidade}_{tipo_atendimento}"[:31]  # Sheet name limited to 31 characters
                            group.to_excel(writer, sheet_name=sheet_name, index=False)

                # Apply conditional formatting to the saved file
                apply_conditional_formatting(output_file)

                st.success(f"SLA Analysis saved to {output_file}")

            except Exception as e:
                st.error(f"An error occurred: {e}")

if __name__ == "__main__":
    main()
